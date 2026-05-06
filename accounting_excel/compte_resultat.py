"""
Module de génération du Compte de Résultat.
Présente les produits et charges avec calcul du résultat net.
"""

from typing import Dict, List
from collections import defaultdict


class CompteResultatManager:
    """
    Gestionnaire de la feuille Compte de Résultat.
    
    Génère le compte de résultat présentant:
    - Les produits (classe 7) par catégorie
    - Les charges (classe 6) par catégorie
    - Le résultat net (Produits - Charges)
    
    Structure selon le PCG français:
    - Résultat d'exploitation
    - Résultat financier
    - Résultat exceptionnel
    - Résultat net
    
    Attributes:
        sheet_name: Nom de la feuille ("Compte de Résultat")
    """
    
    SHEET_NAME = "Compte de Résultat"
    
    COLUMN_WIDTHS = {
        1: 8,    # Niveau/Indent
        2: 50,   # Libellé
        3: 18,   # Montant
    }
    
    # Catégories de produits (classe 7)
    PRODUCTS_CATEGORIES = {
        "70": "Chiffre d'affaires",
        "706": "Prestations de services",
        "707": "Ventes de marchandises",
        "708": "Produits des activités annexes",
        "75": "Autres produits de gestion courante",
        "76": "Produits financiers",
        "77": "Produits exceptionnels",
    }
    
    # Catégories de charges (classe 6)
    CHARGES_CATEGORIES = {
        "60": "Achats",
        "601": "Achats stockés - Matières premières",
        "602": "Achats stockés - Autres approvisionnements",
        "604": "Achats d'études et prestations de services",
        "606": "Achats non stockés",
        "61": "Services extérieurs",
        "611": "Sous-traitance générale",
        "612": "Redevances de crédit-bail",
        "615": "Entretien et réparations",
        "616": "Primes d'assurance",
        "617": "Études et recherches",
        "618": "Divers",
        "62": "Autres services extérieurs",
        "621": "Personnel extérieur à l'entreprise",
        "622": "Rémunérations d'intermédiaires",
        "623": "Publicité et publications",
        "624": "Transports de biens",
        "625": "Déplacements et missions",
        "626": "Frais postaux et télécommunications",
        "627": "Services bancaires",
        "628": "Divers",
        "63": "Impôts et taxes",
        "64": "Charges de personnel",
        "641": "Rémunérations du personnel",
        "645": "Charges de sécurité sociale",
        "647": "Autres charges sociales",
        "648": "Autres charges de personnel",
        "66": "Charges financières",
        "67": "Charges exceptionnelles",
        "68": "Dotations aux amortissements et provisions",
    }
    
    def __init__(self, workbook_builder):
        """
        Initialise le gestionnaire du Compte de Résultat.
        
        Args:
            workbook_builder: Instance de WorkbookBuilder
        """
        self.wb_builder = workbook_builder
        self.styles = workbook_builder.styles
        self.sheet = None
        self._accounts_data = {}
    
    def create_sheet(self) -> None:
        """
        Crée la feuille Compte de Résultat avec sa structure.
        """
        self.sheet = self.wb_builder.workbook.create_sheet(title=self.SHEET_NAME)
        self.wb_builder._sheets[self.SHEET_NAME] = self.sheet
        
        # Titre principal
        title = self.sheet.cell(row=1, column=1, value="COMPTE DE RÉSULTAT")
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=3)
        self.styles.apply_title_style(title)
        self.sheet.row_dimensions[1].height = 25
        
        # En-têtes
        headers = ["Niveau", "Poste", "Montant"]
        for col, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=3, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        self.sheet.freeze_panes = "A4"
    
    def generate_from_entries(self) -> None:
        """
        Génère le Compte de Résultat à partir des écritures.
        
        Extrait uniquement les comptes de classe 6 (charges) et 7 (produits).
        """
        entries = self.wb_builder.get_entries()
        
        if not entries:
            msg_cell = self.sheet.cell(row=4, column=1, 
                                        value="Aucune écriture à afficher")
            msg_cell.font = self.styles.normal_font
            return
        
        # Calculer les totaux par compte (uniquement classes 6 et 7)
        self._accounts_data = self._calculate_account_totals(entries)
        
        # Générer le compte de résultat
        self._generate_statement()
    
    def _calculate_account_totals(self, entries: List[Dict]) -> Dict:
        """
        Calcule les totaux pour les comptes de résultat (classes 6 et 7).
        
        Args:
            entries: Liste des écritures
            
        Returns:
            Dictionnaire {compte: total}
        """
        accounts = defaultdict(float)
        
        for entry in entries:
            account = entry["account"]
            # Uniquement les comptes de résultat
            if not account or account[0] not in ["6", "7"]:
                continue
            
            # Pour les charges (6): débit est positif
            # Pour les produits (7): crédit est positif
            if account[0] == "6":
                accounts[account] += entry["debit"]
            else:  # account[0] == "7"
                accounts[account] += entry["credit"]
        
        return dict(accounts)
    
    def _generate_statement(self) -> None:
        """
        Génère la structure complète du compte de résultat.
        """
        current_row = 4
        
        # === SECTION PRODUITS ===
        row = self._generate_products_section(current_row)
        
        # Ligne total produits
        total_products = sum(v for k, v in self._accounts_data.items() if k[0] == "7")
        row += 1
        total_prod_label = self.sheet.cell(row=row, column=2, value="TOTAL PRODUITS")
        total_prod_label.font = self.styles.total_font
        total_prod_cell = self.sheet.cell(row=row, column=3, value=total_products)
        total_prod_cell.number_format = self.styles.currency_format
        total_prod_cell.font = self.styles.total_font
        total_prod_cell.border = self.styles.thick_border
        self.sheet.row_dimensions[row].height = 20
        
        products_total_row = row
        
        # Espace
        row += 2
        
        # === SECTION CHARGES ===
        row = self._generate_charges_section(row)
        
        # Ligne total charges
        total_charges = sum(v for k, v in self._accounts_data.items() if k[0] == "6")
        row += 1
        total_chg_label = self.sheet.cell(row=row, column=2, value="TOTAL CHARGES")
        total_chg_label.font = self.styles.total_font
        total_chg_cell = self.sheet.cell(row=row, column=3, value=total_charges)
        total_chg_cell.number_format = self.styles.currency_format
        total_chg_cell.font = self.styles.total_font
        total_chg_cell.border = self.styles.thick_border
        self.sheet.row_dimensions[row].height = 20
        
        charges_total_row = row
        
        # Espace
        row += 2
        
        # === RÉSULTAT NET ===
        net_result = total_products - total_charges
        
        # Séparation double
        result_sep_cell = self.sheet.cell(row=row, column=3)
        result_sep_cell.border = self.styles.bottom_border
        
        row += 1
        result_label = self.sheet.cell(row=row, column=2, value="RÉSULTAT NET")
        result_label.font = Font(name='Calibri', size=14, bold=True)
        
        result_cell = self.sheet.cell(row=row, column=3, value=net_result)
        result_cell.number_format = self.styles.currency_format
        result_cell.font = Font(name='Calibri', size=14, bold=True)
        result_cell.fill = self.styles.total_fill
        result_cell.border = self.styles.thick_border
        
        # Colorer selon profit/perte
        if net_result > 0:
            self.styles.apply_positive_style(result_cell)
        elif net_result < 0:
            self.styles.apply_negative_style(result_cell)
        
        self.sheet.row_dimensions[row].height = 25
        
        # Appliquer largeurs de colonnes
        self.apply_column_widths()
    
    def _generate_products_section(self, start_row: int) -> int:
        """
        Génère la section des produits.
        
        Args:
            start_row: Ligne de début
            
        Returns:
            Dernière ligne utilisée
        """
        from openpyxl.styles import Font
        
        row = start_row
        
        # Titre de section
        title = self.sheet.cell(row=row, column=1, value="PRODUITS")
        self.sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        title.fill = self.styles.header_fill
        title.font = self.styles.header_font
        title.alignment = self.styles.center_alignment
        self.sheet.row_dimensions[row].height = 22
        row += 1
        
        # Regrouper par catégorie
        categories = self._group_by_category(
            {k: v for k, v in self._accounts_data.items() if k[0] == "7"}
        )
        
        for category_code in sorted(categories.keys()):
            category_name = self.PRODUCTS_CATEGORIES.get(
                category_code, f"Catégorie {category_code}"
            )
            cat_total = sum(categories[category_code].values())
            
            # Ligne catégorie
            cat_label = self.sheet.cell(row=row, column=2, value=category_name)
            cat_label.font = self.styles.total_font
            cat_cell = self.sheet.cell(row=row, column=3, value=cat_total)
            cat_cell.number_format = self.styles.currency_format
            row += 1
            
            # Détail des comptes (indenté)
            for account in sorted(categories[category_code].keys()):
                amount = categories[category_code][account]
                acc_name = self._get_account_name(account)
                
                acc_label = self.sheet.cell(row=row, column=2, value=f"  {account} - {acc_name}")
                acc_label.alignment = self.styles.left_alignment
                acc_cell = self.sheet.cell(row=row, column=3, value=amount)
                acc_cell.number_format = self.styles.currency_format
                
                # Alternance
                if (row - start_row) % 2 == 0:
                    acc_label.fill = self.styles.alternate_fill
                    acc_cell.fill = self.styles.alternate_fill
                
                row += 1
        
        return row - 1
    
    def _generate_charges_section(self, start_row: int) -> int:
        """
        Génère la section des charges.
        
        Args:
            start_row: Ligne de début
            
        Returns:
            Dernière ligne utilisée
        """
        from openpyxl.styles import Font
        
        row = start_row
        
        # Titre de section
        title = self.sheet.cell(row=row, column=1, value="CHARGES")
        self.sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        title.fill = self.styles.header_fill
        title.font = self.styles.header_font
        title.alignment = self.styles.center_alignment
        self.sheet.row_dimensions[row].height = 22
        row += 1
        
        # Regrouper par catégorie
        categories = self._group_by_category(
            {k: v for k, v in self._accounts_data.items() if k[0] == "6"}
        )
        
        for category_code in sorted(categories.keys()):
            category_name = self.CHARGES_CATEGORIES.get(
                category_code, f"Catégorie {category_code}"
            )
            cat_total = sum(categories[category_code].values())
            
            # Ligne catégorie
            cat_label = self.sheet.cell(row=row, column=2, value=category_name)
            cat_label.font = self.styles.total_font
            cat_cell = self.sheet.cell(row=row, column=3, value=cat_total)
            cat_cell.number_format = self.styles.currency_format
            row += 1
            
            # Détail des comptes (indenté)
            for account in sorted(categories[category_code].keys()):
                amount = categories[category_code][account]
                acc_name = self._get_account_name(account)
                
                acc_label = self.sheet.cell(row=row, column=2, value=f"  {account} - {acc_name}")
                acc_label.alignment = self.styles.left_alignment
                acc_cell = self.sheet.cell(row=row, column=3, value=amount)
                acc_cell.number_format = self.styles.currency_format
                
                # Alternance
                if (row - start_row) % 2 == 0:
                    acc_label.fill = self.styles.alternate_fill
                    acc_cell.fill = self.styles.alternate_fill
                
                row += 1
        
        return row - 1
    
    def _group_by_category(self, accounts: Dict) -> Dict:
        """
        Regroupe les comptes par catégorie (préfixe).
        
        Args:
            accounts: Dictionnaire {compte: montant}
            
        Returns:
            Dictionnaire {catégorie: {compte: montant}}
        """
        categories = defaultdict(dict)
        
        for account, amount in accounts.items():
            # Déterminer le niveau de catégorie
            if len(account) >= 3:
                prefix = account[:3]
            else:
                prefix = account[:2] if len(account) >= 2 else account
            
            categories[prefix][account] = amount
        
        return dict(categories)
    
    def _get_account_name(self, account: str) -> str:
        """
        Récupère l'intitulé d'un compte.
        
        Args:
            account: Numéro du compte
            
        Returns:
            Intitulé du compte
        """
        # Chercher dans les deux dictionnaires
        all_names = {**self.PRODUCTS_CATEGORIES, **self.CHARGES_CATEGORIES}
        
        if account in all_names:
            return all_names[account]
        
        # Essayer par préfixe
        for prefix in [account[:4], account[:3], account[:2]]:
            if prefix in all_names:
                return all_names[prefix]
        
        return f"Compte {account}"
    
    def apply_column_widths(self) -> None:
        """Applique les largeurs de colonnes optimales."""
        from openpyxl.utils import get_column_letter
        for col, width in self.COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col)
            self.sheet.column_dimensions[col_letter].width = width


# Import nécessaire pour Font
from openpyxl.styles import Font
