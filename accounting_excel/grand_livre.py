"""
Module de génération du Grand Livre comptable.
Regroupe toutes les écritures par compte avec soldes intermédiaires.
"""

from datetime import datetime
from typing import List, Dict, Optional
from collections import defaultdict


class GrandLivreManager:
    """
    Gestionnaire de la feuille Grand Livre.
    
    Génère automatiquement le grand livre en regroupant toutes les écritures
    par numéro de compte, avec calcul des soldes intermédiaires et finaux.
    
    Le grand livre présente:
    - Un sommaire avec la liste des comptes
    - Une section détaillée par compte avec toutes ses écritures
    - Les totaux et soldes par compte
    
    Attributes:
        sheet_name: Nom de la feuille ("Grand Livre")
    """
    
    SHEET_NAME = "Grand Livre"
    HEADERS = ["Date", "Journal", "Compte", "Libellé", "Débit", "Crédit", "Solde"]
    
    COLUMN_WIDTHS = {
        1: 12,   # Date
        2: 10,   # Journal
        3: 12,   # Compte
        4: 40,   # Libellé
        5: 14,   # Débit
        6: 14,   # Crédit
        7: 14,   # Solde
    }
    
    def __init__(self, workbook_builder):
        """
        Initialise le gestionnaire du Grand Livre.
        
        Args:
            workbook_builder: Instance de WorkbookBuilder
        """
        self.wb_builder = workbook_builder
        self.styles = workbook_builder.styles
        self.sheet = None
        self._current_row = 1
    
    def create_sheet(self) -> None:
        """
        Crée la feuille Grand Livre avec sa structure.
        Sera remplie après que toutes les écritures soient ajoutées.
        """
        self.sheet = self.wb_builder.workbook.create_sheet(title=self.SHEET_NAME)
        self.wb_builder._sheets[self.SHEET_NAME] = self.sheet
        
        # Titre principal
        title = self.sheet.cell(row=1, column=1, value="GRAND LIVRE COMPTABLE")
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        self.styles.apply_title_style(title)
        self.sheet.row_dimensions[1].height = 25
        
        self._current_row = 3  # Après le titre et une ligne d'espace
    
    def generate_from_entries(self) -> None:
        """
        Génère le Grand Livre à partir de toutes les écritures en mémoire.
        
        Cette méthode:
        1. Regroupe les écritures par compte
        2. Trie les comptes numériquement
        3. Crée une section par compte avec détails et soldes
        """
        entries = self.wb_builder.get_entries()
        
        if not entries:
            # Afficher un message si aucune écriture
            msg_cell = self.sheet.cell(row=self._current_row, column=1, 
                                        value="Aucune écriture à afficher")
            msg_cell.font = self.styles.normal_font
            return
        
        # Regrouper par compte
        accounts_data = defaultdict(list)
        for entry in entries:
            accounts_data[entry["account"]].append(entry)
        
        # Trier les comptes
        sorted_accounts = sorted(accounts_data.keys())
        
        # Créer le sommaire
        self._create_summary(sorted_accounts, accounts_data)
        
        # Générer chaque section de compte
        for account in sorted_accounts:
            self._create_account_section(account, accounts_data[account])
    
    def _create_summary(self, accounts: List[str], accounts_data: Dict) -> None:
        """
        Crée un sommaire avec la liste des comptes et leur nombre d'écritures.
        
        Args:
            accounts: Liste des numéros de comptes triés
            accounts_data: Dictionnaire compte -> écritures
        """
        # Titre du sommaire
        summary_row = self._current_row
        summary_title = self.sheet.cell(row=summary_row, column=1, 
                                         value="SOMMAIRE DES COMPTES")
        summary_title.font = self.styles.total_font
        summary_title.alignment = self.styles.left_alignment
        self.sheet.row_dimensions[summary_row].height = 20
        
        summary_row += 1
        
        # En-têtes du sommaire
        headers = ["Compte", "Intitulé", "Nombre d'écritures", "Total Débit", 
                   "Total Crédit", "Solde"]
        for col, header in enumerate(headers, start=1):
            cell = self.sheet.cell(row=summary_row, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        summary_row += 1
        
        # Lignes du sommaire
        for account in accounts:
            account_entries = accounts_data[account]
            total_debit = sum(e["debit"] for e in account_entries)
            total_credit = sum(e["credit"] for e in account_entries)
            balance = total_debit - total_credit
            
            # Numéro de compte (avec lien vers la section - simulé)
            self.sheet.cell(row=summary_row, column=1, value=account)
            
            # Intitulé (déduit du premier libellé ou générique)
            if account_entries:
                # On pourrait améliorer avec un plan comptable réel
                self.sheet.cell(row=summary_row, column=2, 
                               value=f"Compte {account}")
            
            self.sheet.cell(row=summary_row, column=3, value=len(account_entries))
            
            debit_cell = self.sheet.cell(row=summary_row, column=4, value=total_debit)
            debit_cell.number_format = self.styles.currency_format
            
            credit_cell = self.sheet.cell(row=summary_row, column=5, value=total_credit)
            credit_cell.number_format = self.styles.currency_format
            
            balance_cell = self.sheet.cell(row=summary_row, column=6, value=balance)
            balance_cell.number_format = self.styles.currency_format
            
            # Colorer le solde
            if balance > 0:
                self.styles.apply_positive_style(balance_cell)
            elif balance < 0:
                self.styles.apply_negative_style(balance_cell)
            
            summary_row += 1
        
        # Ligne d'espace avant les détails
        self._current_row = summary_row + 2
    
    def _create_account_section(self, account: str, entries: List[Dict]) -> None:
        """
        Crée la section détaillée pour un compte donné.
        
        Args:
            account: Numéro du compte
            entries: Liste des écritures pour ce compte
        """
        start_row = self._current_row
        
        # Titre du compte
        title = f"COMPTE {account}"
        title_cell = self.sheet.cell(row=start_row, column=1, value=title)
        self.sheet.merge_cells(start_row=start_row, start_column=1, 
                        end_row=start_row, end_column=7)
        self.styles.apply_title_style(title_cell)
        self.sheet.row_dimensions[start_row].height = 22
        
        start_row += 1
        
        # En-têtes du tableau
        for col, header in enumerate(self.HEADERS, start=1):
            cell = self.sheet.cell(row=start_row, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        data_start_row = start_row + 1
        
        # Écritures du compte
        running_balance = 0.0
        for i, entry in enumerate(entries):
            row = data_start_row + i
            
            # Date
            date_val = entry["date"]
            if isinstance(date_val, str):
                try:
                    date_val = datetime.strptime(date_val, "%Y-%m-%d")
                except ValueError:
                    pass
            self.sheet.cell(row=row, column=1, value=date_val).number_format = self.styles.date_format
            
            # Journal
            self.sheet.cell(row=row, column=2, value=entry["journal"])
            
            # Compte
            self.sheet.cell(row=row, column=3, value=account)
            
            # Libellé
            self.sheet.cell(row=row, column=4, value=entry["label"])
            
            # Débit
            debit = entry["debit"]
            debit_cell = self.sheet.cell(row=row, column=5, value=debit if debit else None)
            debit_cell.number_format = self.styles.currency_format
            
            # Crédit
            credit = entry["credit"]
            credit_cell = self.sheet.cell(row=row, column=6, value=credit if credit else None)
            credit_cell.number_format = self.styles.currency_format
            
            # Solde cumulé
            running_balance += debit - credit
            balance_cell = self.sheet.cell(row=row, column=7, value=running_balance)
            balance_cell.number_format = self.styles.currency_format
            
            # Alternance de couleur
            if i % 2 == 0:
                for col in range(1, 8):
                    self.sheet.cell(row=row, column=col).fill = self.styles.alternate_fill
        
        # Totaux du compte
        total_row = data_start_row + len(entries)
        total_debit = sum(e["debit"] for e in entries)
        total_credit = sum(e["credit"] for e in entries)
        final_balance = total_debit - total_credit
        
        # Séparation
        self.sheet.cell(row=total_row - 1, column=5).border = self.styles.thin_border
        self.sheet.cell(row=total_row - 1, column=6).border = self.styles.thin_border
        self.sheet.cell(row=total_row - 1, column=7).border = self.styles.thin_border
        
        # Label TOTAUX
        total_label = self.sheet.cell(row=total_row, column=3, value="TOTAUX")
        total_label.font = self.styles.total_font
        total_label.alignment = self.styles.right_alignment
        
        # Total Débit
        debit_total = self.sheet.cell(row=total_row, column=5, value=total_debit)
        debit_total.number_format = self.styles.currency_format
        debit_total.font = self.styles.total_font
        
        # Total Crédit
        credit_total = self.sheet.cell(row=total_row, column=6, value=total_credit)
        credit_total.number_format = self.styles.currency_format
        credit_total.font = self.styles.total_font
        
        # Solde final
        balance_total = self.sheet.cell(row=total_row, column=7, value=final_balance)
        balance_total.number_format = self.styles.currency_format
        balance_total.font = self.styles.total_font
        balance_total.border = self.styles.thick_border
        
        # Couleur du solde
        if final_balance > 0:
            self.styles.apply_positive_style(balance_total)
        elif final_balance < 0:
            self.styles.apply_negative_style(balance_total)
        
        self.sheet.row_dimensions[total_row].height = 20
        
        # Mettre à jour la position courante
        self._current_row = total_row + 2  # +2 pour espace entre comptes
    
    def apply_column_widths(self) -> None:
        """Applique les largeurs de colonnes optimales."""
        from openpyxl.utils import get_column_letter
        for col, width in self.COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col)
            self.sheet.column_dimensions[col_letter].width = width
        
        # Figer les en-têtes
        self.sheet.freeze_panes = "A4"  # Après titre et en-têtes
