"""
Module de gestion du Journal comptable.
Feuille de saisie des écritures avec validation et mise en forme.
"""

from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import List, Optional
from .utils import AccountingStyles, adjust_column_widths


class JournalManager:
    """
    Gestionnaire de la feuille Journal des écritures comptables.
    
    Crée et gère la feuille principale de saisie avec:
    - Colonnes: Date | Journal | Compte | Libellé | Débit | Crédit
    - Validation de données pour les journaux
    - Formats appropriés (dates, devises)
    - Formules de contrôle d'équilibre
    
    Attributes:
        sheet_name: Nom de la feuille ("Journal")
        headers: Liste des en-têtes de colonnes
    """
    
    SHEET_NAME = "Journal"
    HEADERS = ["Date", "Journal", "Compte", "Libellé", "Débit", "Crédit"]
    
    # Largeurs de colonnes optimales
    COLUMN_WIDTHS = {
        1: 12,   # Date
        2: 10,   # Journal
        3: 12,   # Compte
        4: 45,   # Libellé
        5: 14,   # Débit
        6: 14,   # Crédit
    }
    
    def __init__(self, workbook_builder):
        """
        Initialise le gestionnaire de journal.
        
        Args:
            workbook_builder: Instance de WorkbookBuilder
        """
        self.wb_builder = workbook_builder
        self.styles = workbook_builder.styles
        self.sheet = None
        self._start_row = 2  # Ligne de début des données (après en-têtes)
    
    def create_sheet(self) -> None:
        """
        Crée la feuille Journal avec sa structure complète.
        """
        # Créer la feuille avec en-têtes
        self.sheet = self.wb_builder.create_sheet(
            name=self.SHEET_NAME,
            headers=self.HEADERS,
            column_widths=self.COLUMN_WIDTHS
        )
        
        # Ajouter validation de données pour la colonne Journal (colonne B = 2)
        self.wb_builder.add_data_validation(self.sheet, column=2)
        
        # Ajouter ligne de total avec formules
        self._add_total_row()
        
        # Ajouter titre informatif
        self._add_info_header()
    
    def _add_info_header(self) -> None:
        """Ajoute un en-tête informatif au-dessus du tableau."""
        # Insérer une ligne au début
        self.sheet.insert_rows(1)
        
        # Titre principal
        title_cell = self.sheet.cell(row=1, column=1, value="JOURNAL DES ÉCRITURES COMPTABLES")
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        self.styles.apply_title_style(title_cell)
        self.sheet.row_dimensions[1].height = 25
        
        # Ajuster les références après insertion
        self._start_row = 3  # Les données commencent maintenant à la ligne 3
    
    def _add_total_row(self) -> None:
        """
        Ajoute une ligne de total avec formules Excel dynamiques.
        Utilise SOMME pour calculer les totaux Débit et Crédit.
        """
        # Nous ajouterons cette ligne dynamiquement après les données
        # Pour l'instant, on prépare la structure
        pass
    
    def add_entry(self, date_val, journal: str, account: str,
                  label: str, debit: float, credit: float,
                  row: Optional[int] = None) -> int:
        """
        Ajoute une écriture dans la feuille Journal.
        
        Args:
            date_val: Date de l'écriture (datetime ou string YYYY-MM-DD)
            journal: Code du journal (ACH, VEN, BAN, CAI, etc.)
            account: Numéro de compte (6 chiffres)
            label: Libellé de l'écriture
            debit: Montant au débit
            credit: Montant au crédit
            row: Ligne d'insertion (optionnel, sinon prochaine ligne libre)
            
        Returns:
            Numéro de la ligne où l'écriture a été insérée
        """
        if row is None:
            row = self._get_next_empty_row()
        
        # Formatage de la date
        if isinstance(date_val, str):
            try:
                date_val = datetime.strptime(date_val, "%Y-%m-%d")
            except ValueError:
                pass
        
        # Remplir les cellules
        self.sheet.cell(row=row, column=1, value=date_val).number_format = self.styles.date_format
        self.sheet.cell(row=row, column=2, value=journal.upper())
        self.sheet.cell(row=row, column=3, value=account)
        self.sheet.cell(row=row, column=4, value=label)
        
        # Montants avec format monétaire
        debit_cell = self.sheet.cell(row=row, column=5, value=debit if debit else None)
        debit_cell.number_format = self.styles.currency_format
        
        credit_cell = self.sheet.cell(row=row, column=6, value=credit if credit else None)
        credit_cell.number_format = self.styles.currency_format
        
        # Alternance de couleur des lignes
        if (row - self._start_row) % 2 == 0:
            for col in range(1, 7):
                cell = self.sheet.cell(row=row, column=col)
                cell.fill = self.styles.alternate_fill
        
        return row
    
    def add_entries_from_memory(self) -> None:
        """
        Ajoute toutes les écritures en mémoire dans la feuille Journal.
        """
        entries = self.wb_builder.get_entries()
        for entry in entries:
            self.add_entry(
                date_val=entry["date"],
                journal=entry["journal"],
                account=entry["account"],
                label=entry["label"],
                debit=entry["debit"],
                credit=entry["credit"]
            )
        
        # Ajouter la ligne de total après toutes les écritures
        self._finalize_with_totals()
    
    def _get_next_empty_row(self) -> int:
        """
        Trouve la prochaine ligne vide disponible.
        
        Returns:
            Numéro de la prochaine ligne libre
        """
        max_row = self.sheet.max_row
        if max_row < self._start_row:
            return self._start_row
        
        # Vérifier si la dernière ligne a des données
        last_row_has_data = any(
            self.sheet.cell(row=max_row, column=col).value 
            for col in range(1, 7)
        )
        
        return max_row + 1 if last_row_has_data else max_row
    
    def _finalize_with_totals(self) -> None:
        """
        Ajoute les lignes de totaux et vérification d'équilibre.
        """
        last_data_row = self._get_next_empty_row() - 1
        
        if last_data_row < self._start_row:
            # Aucune donnée, ajouter quand même les totaux
            total_row = self._start_row
        else:
            total_row = last_data_row + 2  # Une ligne d'espace
        
        # Ligne de séparation
        self.sheet.cell(row=total_row - 1, column=5).border = self.styles.thick_border
        self.sheet.cell(row=total_row - 1, column=6).border = self.styles.thick_border
        
        # Label "TOTAUX"
        total_label = self.sheet.cell(row=total_row, column=3, value="TOTAUX")
        total_label.font = self.styles.total_font
        total_label.alignment = self.styles.right_alignment
        
        # Formule SOMME pour le Débit
        debit_formula = f"=SUM(E{self._start_row}:E{last_data_row})"
        debit_cell = self.sheet.cell(row=total_row, column=5, value=debit_formula)
        debit_cell.number_format = self.styles.currency_format
        debit_cell.font = self.styles.total_font
        debit_cell.border = self.styles.thick_border
        
        # Formule SOMME pour le Crédit
        credit_formula = f"=SUM(F{self._start_row}:F{last_data_row})"
        credit_cell = self.sheet.cell(row=total_row, column=6, value=credit_formula)
        credit_cell.number_format = self.styles.currency_format
        credit_cell.font = self.styles.total_font
        credit_cell.border = self.styles.thick_border
        
        # Ligne de vérification d'équilibre
        check_row = total_row + 1
        check_label = self.sheet.cell(row=check_row, column=3, value="ÉQUILIBRE:")
        check_label.font = self.styles.normal_font
        check_label.alignment = self.styles.right_alignment
        
        # Formule de vérification (doit afficher 0 si équilibré)
        check_formula = f"=E{total_row}-F{total_row}"
        check_cell = self.sheet.cell(row=check_row, column=5, value=check_formula)
        check_cell.number_format = self.styles.currency_format
        
        # Message conditionnel (nécessite une formule SI Excel)
        status_formula = f'=IF({check_formula}=0, "✓ ÉQUILIBRÉ", "✗ DÉSÉQUILIBRÉ")'
        status_cell = self.sheet.cell(row=check_row, column=6, value=status_formula)
        
        # Coloration conditionnelle simulée
        if abs(self.wb_builder.total_debit - self.wb_builder.total_credit) < 0.01:
            status_cell.font = self.styles.normal_font  # Vert serait idéal
        else:
            status_cell.font = self.styles.normal_font  # Rouge serait idéal
        
        # Ajuster la hauteur des lignes de total
        self.sheet.row_dimensions[total_row].height = 20
        self.sheet.row_dimensions[check_row].height = 20
    
    def get_entries_count(self) -> int:
        """
        Compte le nombre d'écritures dans le journal.
        
        Returns:
            Nombre d'écritures
        """
        count = 0
        for row in range(self._start_row, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value:  # Si la date existe
                count += 1
        return count
    
    def apply_alternating_colors(self) -> None:
        """Applique un style de lignes alternées sur toutes les données."""
        for row in range(self._start_row, self.sheet.max_row + 1):
            if self.sheet.cell(row=row, column=1).value:  # Si ligne a des données
                if (row - self._start_row) % 2 == 0:
                    for col in range(1, 7):
                        cell = self.sheet.cell(row=row, column=col)
                        if cell.fill.fill_type is None or cell.fill.fill_type == 'none':
                            cell.fill = self.styles.alternate_fill
