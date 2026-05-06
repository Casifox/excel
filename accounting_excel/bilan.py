"""
Module de génération du Bilan comptable.
Présente l'Actif et le Passif avec calcul des totaux et équilibre.
"""

from typing import Dict, List
from collections import defaultdict


class BilanManager:
    """
    Gestionnaire de la feuille Bilan comptable.
    
    Génère le bilan présentant:
    - ACTIF (à gauche): Emplois de l'entreprise
      - Actif immobilisé (classe 2)
      - Actif circulant (classes 3, 4 débiteur, 5)
    - PASSIF (à droite): Ressources de l'entreprise
      - Capitaux propres (classe 1)
      - Dettes (classes 4 créditeur, 5 créditeur)
    
    Le bilan doit toujours être équilibré: Total Actif = Total Passif
    
    Attributes:
        sheet_name: Nom de la feuille ("Bilan")
    """
    
    SHEET_NAME = "Bilan"
    
    # Largeurs de colonnes pour présentation en deux colonnes
    COLUMN_WIDTHS = {
        1: 8,    # Code actif
        2: 35,   # Libellé actif
        3: 18,   # Montant actif
        4: 2,    # Séparateur
        5: 8,    # Code passif
        6: 35,   # Libellé passif
        7: 18,   # Montant passif
    }
    
    # Intitulés des postes d'actif
    ACTIF_LABELS = {
        "2": "IMMOBILISATIONS",
        "20": "Immobilisations incorporelles",
        "21": "Immobilisations corporelles",
        "27": "Immobilisations financières",
        "28": "Amortissements des immobilisations",
        "3": "STOCKS ET EN-COURS",
        "31": "Matières premières",
        "35": "Stocks de produits",
        "37": "Stocks de marchandises",
        "4": "CRÉANCES (Actif Circulant)",
        "40": "Fournisseurs débiteurs",
        "41": "Clients",
        "42": "Personnel et comptes rattachés",
        "44": "État et autres collectivités publiques",
        "46": "Comptes d'attente",
        "5": "TRÉSORERIE - ACTIF",
        "51": "Disponibilités",
        "58": "Virements internes",
    }
    
    # Intitulés des postes de passif
    PASSIF_LABELS = {
        "1": "CAPITAUX PROPRES",
        "10": "Capital et réserves",
        "101": "Capital social",
        "106": "Réserves",
        "11": "Report à nouveau",
        "12": "Résultat de l'exercice",
        "16": "Emprunts et dettes assimilées",
        "4": "DETTES",
        "40": "Fournisseurs et comptes rattachés",
        "41": "Clients créditeurs",
        "42": "Personnel et comptes rattachés",
        "43": "Sécurité sociale et autres organismes sociaux",
        "44": "État et autres collectivités publiques",
        "46": "Comptes d'attente",
        "5": "TRÉSORERIE - PASSIF",
        "51": "Concours bancaires courants",
        "59": "Provisions pour dépréciation",
    }
    
    def __init__(self, workbook_builder):
        """
        Initialise le gestionnaire du Bilan.
        
        Args:
            workbook_builder: Instance de WorkbookBuilder
        """
        self.wb_builder = workbook_builder
        self.styles = workbook_builder.styles
        self.sheet = None
        self._accounts_data = {}
    
    def create_sheet(self) -> None:
        """
        Crée la feuille Bilan avec sa structure en deux colonnes.
        """
        self.sheet = self.wb_builder.workbook.create_sheet(title=self.SHEET_NAME)
        self.wb_builder._sheets[self.SHEET_NAME] = self.sheet
        
        # Titre principal
        title = self.sheet.cell(row=1, column=1, value="BILAN COMPTABLE")
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
        self.styles.apply_title_style(title)
        self.sheet.row_dimensions[1].height = 25
        
        # En-têtes des deux colonnes
        headers_actif = ["Code", "ACTIF", "Montant"]
        headers_passif = ["Code", "PASSIF", "Montant"]
        
        # Colonne ACTIF
        for col, header in enumerate(headers_actif, start=1):
            cell = self.sheet.cell(row=3, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        # Colonne PASSIF (colonnes 5, 6, 7)
        for idx, header in enumerate(headers_passif, start=0):
            col = 5 + idx
            cell = self.sheet.cell(row=3, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        # Colonne séparatrice
        self.sheet.column_dimensions["D"].width = 2
        
        self.sheet.freeze_panes = "A4"
    
    def generate_from_entries(self) -> None:
        """
        Génère le Bilan à partir des écritures.
        
        Calcule les soldes de tous les comptes de bilan (classes 1-5)
        et les répartit entre Actif et Passif.
        """
        entries = self.wb_builder.get_entries()
        
        if not entries:
            msg_cell = self.sheet.cell(row=4, column=1, 
                                        value="Aucune écriture à afficher")
            msg_cell.font = self.styles.normal_font
            return
        
        # Calculer les soldes par compte
        self._accounts_data = self._calculate_account_balances(entries)
        
        # Générer le bilan
        self._generate_balance_sheet()
    
    def _calculate_account_balances(self, entries: List[Dict]) -> Dict:
        """
        Calcule les soldes pour les comptes de bilan (classes 1-5).
        
        Args:
            entries: Liste des écritures
            
        Returns:
            Dictionnaire {compte: {debit, credit, solde, type}}
        """
        accounts = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
        
        for entry in entries:
            account = entry["account"]
            # Uniquement les comptes de bilan
            if not account or account[0] not in ["1", "2", "3", "4", "5"]:
                continue
            
            accounts[account]["debit"] += entry["debit"]
            accounts[account]["credit"] += entry["credit"]
        
        # Calculer les soldes et déterminer nature (actif/passif)
        result = {}
        for account, data in accounts.items():
            balance = data["debit"] - data["credit"]
            
            # Déterminer si c'est un actif ou passif selon la classe
            class_code = account[0]
            if class_code in ["1", "2", "3", "4", "5"]:
                # Par défaut, solde débiteur = actif, créditeur = passif
                # Mais certaines classes ont une nature prédéfinie
                if class_code == "2":  # Immobilisations (toujours à l'actif)
                    account_type = "actif"
                elif class_code == "1":  # Capitaux (toujours au passif)
                    account_type = "passif"
                elif class_code == "3":  # Stocks (actif si positif)
                    account_type = "actif" if balance >= 0 else "passif"
                elif class_code == "4":  # Tiers (dépend du solde)
                    account_type = "actif" if balance >= 0 else "passif"
                elif class_code == "5":  # Financiers (dépend du solde)
                    account_type = "actif" if balance >= 0 else "passif"
                else:
                    account_type = "actif" if balance >= 0 else "passif"
                
                result[account] = {
                    "debit": data["debit"],
                    "credit": data["credit"],
                    "balance": abs(balance),
                    "type": account_type,
                    "is_null": abs(balance) < 0.01
                }
        
        return result
    
    def _generate_balance_sheet(self) -> None:
        """
        Génère la structure complète du bilan en deux colonnes.
        """
        # Séparer les comptes en actif et passif
        actif_accounts = {k: v for k, v in self._accounts_data.items() 
                         if v["type"] == "actif" and not v["is_null"]}
        passif_accounts = {k: v for k, v in self._accounts_data.items() 
                          if v["type"] == "passif" and not v["is_null"]}
        
        # Générer chaque colonne
        actif_row = self._generate_actif_section(actif_accounts, start_row=4)
        passif_row = self._generate_passif_section(passif_accounts, start_row=4)
        
        # Ajouter les totaux
        total_actif = sum(v["balance"] for v in actif_accounts.values())
        total_passif = sum(v["balance"] for v in passif_accounts.values())
        
        # Ligne de totaux
        max_rows = max(actif_row, passif_row)
        
        # Séparation double
        for col in [3, 7]:
            sep_cell = self.sheet.cell(row=max_rows, column=col)
            sep_cell.border = self.styles.bottom_border
        
        # Totaux
        total_actif_label = self.sheet.cell(row=max_rows + 1, column=2, 
                                            value="TOTAL ACTIF")
        total_actif_label.font = self.styles.total_font
        total_actif_label.alignment = self.styles.right_alignment
        
        total_actif_cell = self.sheet.cell(row=max_rows + 1, column=3, 
                                           value=total_actif)
        total_actif_cell.number_format = self.styles.currency_format
        total_actif_cell.font = self.styles.total_font
        total_actif_cell.border = self.styles.thick_border
        
        total_passif_label = self.sheet.cell(row=max_rows + 1, column=6, 
                                             value="TOTAL PASSIF")
        total_passif_label.font = self.styles.total_font
        total_passif_label.alignment = self.styles.right_alignment
        
        total_passif_cell = self.sheet.cell(row=max_rows + 1, column=7, 
                                            value=total_passif)
        total_passif_cell.number_format = self.styles.currency_format
        total_passif_cell.font = self.styles.total_font
        total_passif_cell.border = self.styles.thick_border
        
        self.sheet.row_dimensions[max_rows + 1].height = 22
        
        # Vérification d'équilibre
        check_row = max_rows + 2
        if abs(total_actif - total_passif) < 0.01:
            status = "✓ BILAN ÉQUILIBRÉ"
        else:
            diff = total_actif - total_passif
            status = f"✗ ÉCART: {diff:,.2f} €"
        
        status_cell = self.sheet.cell(row=check_row, column=3, value=status)
        self.sheet.merge_cells(start_row=check_row, start_column=3, 
                         end_row=check_row, end_column=5)
        
        # Appliquer largeurs
        self.apply_column_widths()
    
    def _generate_actif_section(self, accounts: Dict, start_row: int) -> int:
        """
        Génère la section Actif du bilan.
        
        Args:
            accounts: Dictionnaire des comptes d'actif
            start_row: Ligne de début
            
        Returns:
            Dernière ligne utilisée
        """
        row = start_row
        
        # Titre de section
        title = self.sheet.cell(row=row, column=1, value="ACTIF")
        self.sheet.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
        title.fill = self.styles.header_fill
        title.font = self.styles.header_font
        title.alignment = self.styles.center_alignment
        self.sheet.row_dimensions[row].height = 22
        row += 1
        
        # Regrouper par classe
        classes = self._group_by_class(accounts)
        
        for class_code in sorted(classes.keys()):
            class_accounts = classes[class_code]
            class_label = self.ACTIF_LABELS.get(class_code, f"Classe {class_code}")
            
            # Titre de classe
            class_cell = self.sheet.cell(row=row, column=2, value=class_label)
            class_cell.font = self.styles.total_font
            class_cell.border = self.styles.thin_border
            row += 1
            
            # Comptes de la classe
            for account in sorted(class_accounts.keys()):
                data = class_accounts[account]
                acc_label = self._get_account_label(account, self.ACTIF_LABELS)
                
                self.sheet.cell(row=row, column=1, value=account)
                label_cell = self.sheet.cell(row=row, column=2, value=acc_label)
                label_cell.alignment = self.styles.left_alignment
                
                amount_cell = self.sheet.cell(row=row, column=3, value=data["balance"])
                amount_cell.number_format = self.styles.currency_format
                
                # Alternance
                if (row - start_row) % 2 == 1:
                    for col in range(1, 4):
                        self.sheet.cell(row=row, column=col).fill = self.styles.alternate_fill
                
                row += 1
            
            # Espace entre classes
            row += 1
        
        return row - 1
    
    def _generate_passif_section(self, accounts: Dict, start_row: int) -> int:
        """
        Génère la section Passif du bilan.
        
        Args:
            accounts: Dictionnaire des comptes de passif
            start_row: Ligne de début
            
        Returns:
            Dernière ligne utilisée
        """
        row = start_row
        
        # Titre de section
        title = self.sheet.cell(row=row, column=5, value="PASSIF")
        self.sheet.merge_cells(start_row=row, start_column=5, end_row=row, end_column=7)
        title.fill = self.styles.header_fill
        title.font = self.styles.header_font
        title.alignment = self.styles.center_alignment
        self.sheet.row_dimensions[row].height = 22
        row += 1
        
        # Regrouper par classe
        classes = self._group_by_class(accounts)
        
        for class_code in sorted(classes.keys()):
            class_accounts = classes[class_code]
            class_label = self.PASSIF_LABELS.get(class_code, f"Classe {class_code}")
            
            # Titre de classe
            class_cell = self.sheet.cell(row=row, column=6, value=class_label)
            class_cell.font = self.styles.total_font
            class_cell.border = self.styles.thin_border
            row += 1
            
            # Comptes de la classe
            for account in sorted(class_accounts.keys()):
                data = class_accounts[account]
                acc_label = self._get_account_label(account, self.PASSIF_LABELS)
                
                self.sheet.cell(row=row, column=5, value=account)
                label_cell = self.sheet.cell(row=row, column=6, value=acc_label)
                label_cell.alignment = self.styles.left_alignment
                
                amount_cell = self.sheet.cell(row=row, column=7, value=data["balance"])
                amount_cell.number_format = self.styles.currency_format
                
                # Alternance
                if (row - start_row) % 2 == 1:
                    for col in range(5, 8):
                        self.sheet.cell(row=row, column=col).fill = self.styles.alternate_fill
                
                row += 1
            
            # Espace entre classes
            row += 1
        
        return row - 1
    
    def _group_by_class(self, accounts: Dict) -> Dict:
        """
        Regroupe les comptes par classe (premier chiffre).
        
        Args:
            accounts: Dictionnaire des comptes
            
        Returns:
            Dictionnaire {classe: {compte: données}}
        """
        classes = defaultdict(dict)
        
        for account, data in accounts.items():
            class_code = account[0] if len(account) > 0 else "0"
            classes[class_code][account] = data
        
        return dict(classes)
    
    def _get_account_label(self, account: str, labels_dict: Dict) -> str:
        """
        Récupère l'intitulé d'un compte.
        
        Args:
            account: Numéro du compte
            labels_dict: Dictionnaire des intitulés
            
        Returns:
            Intitulé du compte
        """
        # Chercher correspondance exacte
        if account in labels_dict:
            return labels_dict[account]
        
        # Chercher par préfixe décroissant
        for length in [4, 3, 2, 1]:
            prefix = account[:length]
            if prefix in labels_dict:
                return labels_dict[prefix]
        
        return f"Compte {account}"
    
    def apply_column_widths(self) -> None:
        """Applique les largeurs de colonnes optimales."""
        from openpyxl.utils import get_column_letter
        for col, width in self.COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col)
            self.sheet.column_dimensions[col_letter].width = width
