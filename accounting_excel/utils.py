"""
Module utilitaire pour les styles Excel, formats et fonctions helpers.
Gère la mise en forme professionnelle des fichiers comptables.
"""

from openpyxl.styles import (
    Font, Fill, PatternFill, Border, Side, Alignment, Color
)
from openpyxl.utils import get_column_letter
from typing import Optional, Tuple
from datetime import datetime


# Couleurs professionnelles pour la comptabilité
COLORS = {
    "header_bg": "4472C4",      # Bleu professionnel pour en-têtes
    "header_font": "FFFFFF",     # Blanc pour texte en-têtes
    "title_bg": "D9E1F2",        # Bleu clair pour titres
    "total_bg": "FFC000",        # Jaune/orange pour totaux
    "positive": "00B050",        # Vert pour positifs/crédit
    "negative": "FF0000",        # Rouge pour négatifs/débit
    "border": "000000",          # Noir pour bordures
    "alternate_row": "F2F2F2",   # Gris clair pour lignes alternées
}


class AccountingStyles:
    """
    Classe gérant tous les styles Excel utilisés dans les documents comptables.
    
    Attributes:
        header_font: Police pour les en-têtes
        header_fill: Remplissage pour les en-têtes
        normal_font: Police normale
        currency_format: Format monétaire EUR
        date_format: Format de date français
        title_alignment: Alignement pour les titres
    """
    
    def __init__(self):
        """Initialise tous les styles prédéfinis."""
        # Bordures
        self.thin_border = Border(
            left=Side(style='thin', color=COLORS["border"]),
            right=Side(style='thin', color=COLORS["border"]),
            top=Side(style='thin', color=COLORS["border"]),
            bottom=Side(style='thin', color=COLORS["border"])
        )
        
        self.thick_border = Border(
            left=Side(style='thick', color=COLORS["border"]),
            right=Side(style='thick', color=COLORS["border"]),
            top=Side(style='thick', color=COLORS["border"]),
            bottom=Side(style='thick', color=COLORS["border"])
        )
        
        self.bottom_border = Border(
            bottom=Side(style='double', color=COLORS["border"])
        )
        
        # Fonts
        self.header_font = Font(
            name='Calibri',
            size=11,
            bold=True,
            color=COLORS["header_font"]
        )
        
        self.title_font = Font(
            name='Calibri',
            size=14,
            bold=True,
            color="000000"
        )
        
        self.normal_font = Font(
            name='Calibri',
            size=11
        )
        
        self.total_font = Font(
            name='Calibri',
            size=11,
            bold=True
        )
        
        # Fills
        self.header_fill = PatternFill(
            start_color=COLORS["header_bg"],
            end_color=COLORS["header_bg"],
            fill_type='solid'
        )
        
        self.title_fill = PatternFill(
            start_color=COLORS["title_bg"],
            end_color=COLORS["title_bg"],
            fill_type='solid'
        )
        
        self.total_fill = PatternFill(
            start_color=COLORS["total_bg"],
            end_color=COLORS["total_bg"],
            fill_type='solid'
        )
        
        self.alternate_fill = PatternFill(
            start_color=COLORS["alternate_row"],
            end_color=COLORS["alternate_row"],
            fill_type='solid'
        )
        
        # Alignments
        self.center_alignment = Alignment(
            horizontal='center',
            vertical='center',
            wrap_text=False
        )
        
        self.left_alignment = Alignment(
            horizontal='left',
            vertical='center'
        )
        
        self.right_alignment = Alignment(
            horizontal='right',
            vertical='center'
        )
        
        # Formats
        self.currency_format = '# ##0.00 €'
        self.date_format = 'DD/MM/YYYY'
        self.account_format = '000000'
    
    def apply_header_style(self, cell):
        """Applique le style d'en-tête à une cellule."""
        cell.font = self.header_font
        cell.fill = self.header_fill
        cell.alignment = self.center_alignment
        cell.border = self.thin_border
        return cell
    
    def apply_title_style(self, cell):
        """Applique le style de titre à une cellule."""
        cell.font = self.title_font
        cell.fill = self.title_fill
        cell.alignment = self.center_alignment
        cell.border = self.thin_border
        return cell
    
    def apply_currency_format(self, cell):
        """Applique le format monétaire à une cellule."""
        cell.number_format = self.currency_format
        return cell
    
    def apply_date_format(self, cell):
        """Applique le format de date à une cellule."""
        cell.number_format = self.date_format
        return cell
    
    def apply_positive_style(self, cell):
        """Applique le style vert pour valeurs positives."""
        cell.font = Font(color=COLORS["positive"], bold=True)
        return cell
    
    def apply_negative_style(self, cell):
        """Applique le style rouge pour valeurs négatives."""
        cell.font = Font(color=COLORS["negative"], bold=True)
        return cell


def format_currency(value: float, symbol: str = "€") -> str:
    """
    Formate un nombre en devise française.
    
    Args:
        value: La valeur numérique à formater
        symbol: Le symbole de devise (par défaut: €)
    
    Returns:
        Chaîne formatée avec séparateurs de milliers et décimales
    
    Example:
        >>> format_currency(1234567.89)
        '1 234 567,89 €'
    """
    if value is None:
        value = 0.0
    return f"{value:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + f" {symbol}"


def validate_entry(date_val, journal: str, account: str, 
                   debit: float, credit: float) -> Tuple[bool, str]:
    """
    Valide une écriture comptable avant insertion.
    
    Args:
        date_val: Date de l'écriture (datetime ou string)
        journal: Code du journal (ACH, VEN, BAN, CAI)
        account: Numéro de compte (6 chiffres)
        debit: Montant au débit
        credit: Montant au crédit
    
    Returns:
        Tuple (is_valid, error_message)
    """
    valid_journals = ["ACH", "VEN", "BAN", "CAI", "OD", "Rappro"]
    
    # Validation date
    if date_val is None:
        return False, "La date est obligatoire"
    
    try:
        if isinstance(date_val, str):
            datetime.strptime(date_val, "%Y-%m-%d")
        elif not isinstance(date_val, datetime):
            return False, "Format de date invalide"
    except ValueError:
        return False, "Format de date invalide (attendu: YYYY-MM-DD)"
    
    # Validation journal
    if journal not in valid_journals:
        return False, f"Journal invalide. Doit être parmi: {', '.join(valid_journals)}"
    
    # Validation compte (6 chiffres)
    if not account or not account.isdigit() or len(account) != 6:
        return False, "Le numéro de compte doit comporter 6 chiffres"
    
    # Validation montants
    try:
        debit = float(debit) if debit else 0.0
        credit = float(credit) if credit else 0.0
    except (ValueError, TypeError):
        return False, "Les montants doivent être numériques"
    
    if debit < 0 or credit < 0:
        return False, "Les montants ne peuvent pas être négatifs"
    
    if debit == 0 and credit == 0:
        return False, "Au moins un montant (débit ou crédit) doit être renseigné"
    
    if debit > 0 and credit > 0:
        return False, "Une écriture ne peut pas avoir à la fois un débit et un crédit"
    
    return True, ""


def adjust_column_widths(worksheet, column_widths: dict):
    """
    Ajuste automatiquement la largeur des colonnes.
    
    Args:
        worksheet: Feuille Excel openpyxl
        column_widths: Dictionnaire {colonne: largeur}
    """
    for col, width in column_widths.items():
        if isinstance(col, int):
            column_letter = get_column_letter(col)
        else:
            column_letter = col
        worksheet.column_dimensions[column_letter].width = width


def get_account_class(account: str) -> str:
    """
    Détermine la classe d'un compte comptable selon le PCG français.
    
    Args:
        account: Numéro de compte (6 chiffres)
    
    Returns:
        Nom de la classe du compte
    """
    if not account or len(account) < 1:
        return "Inconnu"
    
    first_digit = account[0]
    classes = {
        "1": "Comptes de capitaux",
        "2": "Comptes d'immobilisations",
        "3": "Comptes de stocks",
        "4": "Comptes de tiers",
        "5": "Comptes financiers",
        "6": "Comptes de charges",
        "7": "Comptes de produits",
        "8": "Comptes spéciaux",
    }
    
    return classes.get(first_digit, "Compte hors bilan")


def is_balance_sheet_account(account: str) -> bool:
    """
    Vérifie si un compte est un compte de bilan.
    
    Args:
        account: Numéro de compte
    
    Returns:
        True si compte de bilan (classes 1-5), False sinon
    """
    if not account or len(account) < 1:
        return False
    return account[0] in ["1", "2", "3", "4", "5"]


def is_income_statement_account(account: str) -> bool:
    """
    Vérifie si un compte est un compte de résultat.
    
    Args:
        account: Numéro de compte
    
    Returns:
        True si compte de résultat (classes 6-7), False sinon
    """
    if not account or len(account) < 1:
        return False
    return account[0] in ["6", "7"]
