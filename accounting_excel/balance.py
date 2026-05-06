"""
Module de génération de la Balance comptable.
Présente tous les comptes avec totaux débit, crédit et soldes.
"""

from typing import Dict, List
from collections import defaultdict


class BalanceManager:
    """
    Gestionnaire de la feuille Balance comptable.
    
    Génère la balance générale présentant pour chaque compte:
    - Total des mouvements Débit
    - Total des mouvements Crédit
    - Solde débiteur ou créditeur
    
    La balance est organisée par classes de comptes selon le PCG français:
    - Classe 1: Capitaux
    - Classe 2: Immobilisations
    - Classe 3: Stocks
    - Classe 4: Tiers
    - Classe 5: Financiers
    - Classe 6: Charges
    - Classe 7: Produits
    
    Attributes:
        sheet_name: Nom de la feuille ("Balance")
    """
    
    SHEET_NAME = "Balance"
    HEADERS = ["Compte", "Intitulé", "Débit", "Crédit", "Solde Débiteur", "Solde Créditeur"]
    
    COLUMN_WIDTHS = {
        1: 14,   # Compte
        2: 45,   # Intitulé
        3: 16,   # Débit
        4: 16,   # Crédit
        5: 18,   # Solde Débiteur
        6: 18,   # Solde Créditeur
    }
    
    # Plan comptable simplifié (peut être étendu)
    ACCOUNT_NAMES = {
        "101": "Capital social",
        "106": "Réserves",
        "11": "Report à nouveau",
        "12": "Résultat de l'exercice",
        "20": "Immobilisations incorporelles",
        "21": "Immobilisations corporelles",
        "28": "Amortissements",
        "401": "Fournisseurs",
        "409": "Fournisseurs débiteurs",
        "411": "Clients",
        "419": "Clients créditeurs",
        "421": "Personnel - Rémunérations dues",
        "431": "Sécurité sociale",
        "445": "TVA",
        "467": "Comptes d'attente",
        "512": "Banque",
        "514": "CCP",
        "519": "Emprunts bancaires",
        "58": "Virements internes",
        "60": "Achats",
        "61": "Services extérieurs",
        "62": "Autres services extérieurs",
        "63": "Impôts et taxes",
        "64": "Charges de personnel",
        "66": "Charges financières",
        "67": "Charges exceptionnelles",
        "68": "Dotations aux amortissements",
        "70": "Ventes",
        "75": "Produits divers",
        "76": "Produits financiers",
        "77": "Produits exceptionnels",
    }
    
    def __init__(self, workbook_builder):
        """
        Initialise le gestionnaire de la Balance.
        
        Args:
            workbook_builder: Instance de WorkbookBuilder
        """
        self.wb_builder = workbook_builder
        self.styles = workbook_builder.styles
        self.sheet = None
        self._accounts_data = {}
    
    def create_sheet(self) -> None:
        """
        Crée la feuille Balance avec sa structure.
        Sera remplie après que toutes les écritures soient traitées.
        """
        self.sheet = self.wb_builder.workbook.create_sheet(title=self.SHEET_NAME)
        self.wb_builder._sheets[self.SHEET_NAME] = self.sheet
        
        # Titre principal
        title = self.sheet.cell(row=1, column=1, value="BALANCE COMPTABLE")
        self.sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
        self.styles.apply_title_style(title)
        self.sheet.row_dimensions[1].height = 25
        
        # Sous-titre avec informations
        subtitle = "Arrêté au " + "(date de clôture)"
        subtitle_cell = self.sheet.cell(row=2, column=1, value=subtitle)
        subtitle_cell.font = self.styles.normal_font
        subtitle_cell.alignment = self.styles.left_alignment
        self.sheet.row_dimensions[2].height = 18
    
    def generate_from_entries(self) -> None:
        """
        Génère la Balance à partir de toutes les écritures en mémoire.
        
        Calcule pour chaque compte:
        - Total Débit
        - Total Crédit
        - Solde (débiteur ou créditeur)
        """
        entries = self.wb_builder.get_entries()
        
        if not entries:
            msg_cell = self.sheet.cell(row=4, column=1, 
                                        value="Aucune écriture à afficher")
            msg_cell.font = self.styles.normal_font
            return
        
        # Calculer les totaux par compte
        self._accounts_data = self._calculate_account_totals(entries)
        
        # Ajouter les en-têtes
        header_row = 4
        for col, header in enumerate(self.HEADERS, start=1):
            cell = self.sheet.cell(row=header_row, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        # Générer les lignes par classe de compte
        self._generate_balance_lines()
        
        # Ajouter les totaux généraux
        self._add_grand_totals()
    
    def _calculate_account_totals(self, entries: List[Dict]) -> Dict:
        """
        Calcule les totaux débit/crédit pour chaque compte.
        
        Args:
            entries: Liste des écritures
            
        Returns:
            Dictionnaire {compte: {debit, credit, solde}}
        """
        accounts = defaultdict(lambda: {"debit": 0.0, "credit": 0.0})
        
        for entry in entries:
            account = entry["account"]
            accounts[account]["debit"] += entry["debit"]
            accounts[account]["credit"] += entry["credit"]
        
        # Calculer les soldes
        result = {}
        for account, data in accounts.items():
            balance = data["debit"] - data["credit"]
            result[account] = {
                "debit": data["debit"],
                "credit": data["credit"],
                "balance": balance,
                "is_debitor": balance >= 0
            }
        
        return result
    
    def _generate_balance_lines(self) -> None:
        """
        Génère les lignes de la balance triées par classe de compte.
        """
        # Trier les comptes
        sorted_accounts = sorted(self._accounts_data.keys())
        
        # Regrouper par classe
        classes = defaultdict(list)
        for account in sorted_accounts:
            class_code = account[0] if len(account) > 0 else "0"
            classes[class_code].append(account)
        
        current_row = 5
        grand_total_debit = 0.0
        grand_total_credit = 0.0
        
        class_names = {
            "1": "CLASSE 1 - COMPTES DE CAPITAUX",
            "2": "CLASSE 2 - COMPTES D'IMMOBILISATIONS",
            "3": "CLASSE 3 - COMPTES DE STOCKS",
            "4": "CLASSE 4 - COMPTES DE TIERS",
            "5": "CLASSE 5 - COMPTES FINANCIERS",
            "6": "CLASSE 6 - COMPTES DE CHARGES",
            "7": "CLASSE 7 - COMPTES DE PRODUITS",
            "8": "CLASSE 8 - COMPTES SPÉCIAUX",
        }
        
        for class_code in sorted(classes.keys()):
            class_accounts = classes[class_code]
            
            if not class_accounts:
                continue
            
            # Titre de classe
            class_name = class_names.get(class_code, f"CLASSE {class_code}")
            class_cell = self.sheet.cell(row=current_row, column=1, value=class_name)
            self.sheet.merge_cells(start_row=current_row, start_column=1, 
                           end_row=current_row, end_column=6)
            class_cell.fill = self.styles.title_fill
            class_cell.font = self.styles.total_font
            class_cell.border = self.styles.thin_border
            self.sheet.row_dimensions[current_row].height = 22
            current_row += 1
            
            # Comptes de la classe
            for account in class_accounts:
                data = self._accounts_data[account]
                
                # Numéro de compte
                self.sheet.cell(row=current_row, column=1, value=account)
                
                # Intitulé
                intitule = self._get_account_name(account)
                self.sheet.cell(row=current_row, column=2, value=intitule)
                
                # Total Débit
                debit_cell = self.sheet.cell(row=current_row, column=3, value=data["debit"])
                debit_cell.number_format = self.styles.currency_format
                
                # Total Crédit
                credit_cell = self.sheet.cell(row=current_row, column=4, value=data["credit"])
                credit_cell.number_format = self.styles.currency_format
                
                # Solde (dans la colonne appropriée)
                if data["is_debitor"]:
                    # Solde débiteur
                    balance_cell = self.sheet.cell(row=current_row, column=5, 
                                                   value=abs(data["balance"]))
                    balance_cell.number_format = self.styles.currency_format
                    if data["balance"] > 0:
                        self.styles.apply_positive_style(balance_cell)
                else:
                    # Solde créditeur
                    balance_cell = self.sheet.cell(row=current_row, column=6, 
                                                   value=abs(data["balance"]))
                    balance_cell.number_format = self.styles.currency_format
                    self.styles.apply_negative_style(balance_cell)
                
                # Alternance de couleur
                row_index = current_row - 5
                if row_index % 2 == 0:
                    for col in range(1, 7):
                        self.sheet.cell(row=current_row, column=col).fill = self.styles.alternate_fill
                
                current_row += 1
                
                # Accumuler les totaux
                grand_total_debit += data["debit"]
                grand_total_credit += data["credit"]
            
            # Ligne d'espace entre classes
            current_row += 1
        
        self._final_row = current_row
        self._grand_total_debit = grand_total_debit
        self._grand_total_credit = grand_total_credit
    
    def _get_account_name(self, account: str) -> str:
        """
        Récupère l'intitulé d'un compte depuis le plan comptable.
        
        Args:
            account: Numéro du compte
            
        Returns:
            Intitulé du compte ou générique si inconnu
        """
        # Essayer correspondance exacte
        if account in self.ACCOUNT_NAMES:
            return self.ACCOUNT_NAMES[account]
        
        # Essair correspondance par préfixe (classe/compte principal)
        for prefix in [account[:4], account[:3], account[:2], account[:1]]:
            for acc_name, name in self.ACCOUNT_NAMES.items():
                if acc_name.startswith(prefix) or prefix.startswith(acc_name[:len(prefix)]):
                    if prefix in self.ACCOUNT_NAMES:
                        return self.ACCOUNT_NAMES[prefix]
        
        return f"Compte {account}"
    
    def _add_grand_totals(self) -> None:
        """
        Ajoute les lignes de totaux généraux de la balance.
        """
        total_row = self._final_row
        
        # Séparation double
        for col in [3, 4, 5, 6]:
            cell = self.sheet.cell(row=total_row - 1, column=col)
            cell.border = self.styles.bottom_border
        
        # Label TOTAUX
        total_label = self.sheet.cell(row=total_row, column=2, value="TOTAUX GÉNÉRAUX")
        total_label.font = self.styles.total_font
        total_label.alignment = self.styles.right_alignment
        
        # Total Débit général
        debit_total = self.sheet.cell(row=total_row, column=3, value=self._grand_total_debit)
        debit_total.number_format = self.styles.currency_format
        debit_total.font = self.styles.total_font
        debit_total.border = self.styles.thick_border
        
        # Total Crédit général
        credit_total = self.sheet.cell(row=total_row, column=4, value=self._grand_total_credit)
        credit_total.number_format = self.styles.currency_format
        credit_total.font = self.styles.total_font
        credit_total.border = self.styles.thick_border
        
        # Vérification équilibre
        balance_check = self._grand_total_debit - self._grand_total_credit
        
        if abs(balance_check) < 0.01:
            status = "✓ BALANCE ÉQUILIBRÉE"
            status_color = self.styles.normal_font
        else:
            status = f"✗ ÉCART: {balance_check:,.2f} €"
            status_color = self.styles.normal_font
        
        status_cell = self.sheet.cell(row=total_row, column=5, value=status)
        self.sheet.merge_cells(start_row=total_row, start_column=5, 
                         end_row=total_row, end_column=6)
        status_cell.font = status_color
        status_cell.border = self.styles.thick_border
        
        self.sheet.row_dimensions[total_row].height = 22
        
        # Figer les en-têtes
        self.sheet.freeze_panes = "A5"
    
    def apply_column_widths(self) -> None:
        """Applique les largeurs de colonnes optimales."""
        from openpyxl.utils import get_column_letter
        for col, width in self.COLUMN_WIDTHS.items():
            col_letter = get_column_letter(col)
            self.sheet.column_dimensions[col_letter].width = width
