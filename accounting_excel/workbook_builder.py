"""
Module de création et gestion du classeur Excel principal.
Gère la structure globale du fichier et les feuilles de calcul.
"""

from openpyxl import Workbook
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from typing import Optional, List
from .utils import AccountingStyles, adjust_column_widths


class WorkbookBuilder:
    """
    Classe principale pour créer et gérer le classeur Excel comptable.
    
    Attributes:
        workbook: Instance du classeur openpyxl
        styles: Instance des styles comptables
        filename: Nom du fichier Excel
    """
    
    # Journaux comptables standards
    VALID_JOURNALS = ["ACH", "VEN", "BAN", "CAI", "OD", "Rappro"]
    
    def __init__(self, filename: str):
        """
        Initialise le constructeur de classeur.
        
        Args:
            filename: Chemin complet du fichier Excel à créer
        """
        self.filename = filename
        self.workbook = Workbook()
        self.styles = AccountingStyles()
        self._sheets = {}
        self._entries = []  # Stockage des écritures en mémoire
        
        # Supprimer la feuille par défaut
        if "Sheet" in self.workbook.sheetnames:
            del self.workbook["Sheet"]
    
    def create_sheet(self, name: str, headers: List[str], 
                     column_widths: Optional[dict] = None) -> None:
        """
        Crée une nouvelle feuille avec des en-têtes prédéfinis.
        
        Args:
            name: Nom de la feuille
            headers: Liste des noms de colonnes
            column_widths: Dictionnaire optionnel {colonne: largeur}
        """
        ws = self.workbook.create_sheet(title=name)
        self._sheets[name] = ws
        
        # Ajouter les en-têtes
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            self.styles.apply_header_style(cell)
        
        # Ajuster les largeurs de colonnes
        if column_widths:
            adjust_column_widths(ws, column_widths)
        
        # Figer la première ligne
        ws.freeze_panes = "A2"
        
        return ws
    
    def add_data_validation(self, worksheet, column: int, 
                           validation_type: str = "list",
                           formula: str = None) -> None:
        """
        Ajoute une validation de données à une colonne.
        
        Args:
            worksheet: Feuille cible
            column: Numéro de colonne (1-based)
            validation_type: Type de validation ("list", "whole", "decimal")
            formula: Formule ou liste pour la validation
        """
        if validation_type == "list":
            if formula is None:
                formula = '"' + ",".join(self.VALID_JOURNALS) + '"'
            
            dv = DataValidation(
                type="list",
                formula1=formula,
                allow_blank=True,
                showDropDown=False  # True cache la liste déroulante dans Excel
            )
            dv.error = "Veuillez sélectionner un journal valide"
            dv.errorTitle = "Journal invalide"
            
            # Appliquer à la colonne entière (sauf en-tête)
            col_letter = get_column_letter(column)
            dv.add(f"{col_letter}2:{col_letter}1048576")
            worksheet.add_data_validation(dv)
    
    def get_sheet(self, name: str):
        """
        Récupère une feuille existante.
        
        Args:
            name: Nom de la feuille
            
        Returns:
            La feuille demandée ou None si elle n'existe pas
        """
        return self._sheets.get(name)
    
    def add_entry_to_memory(self, date_val, journal: str, account: str,
                            label: str, debit: float, credit: float) -> dict:
        """
        Ajoute une écriture en mémoire pour traitement ultérieur.
        
        Args:
            date_val: Date de l'écriture
            journal: Code journal
            account: Numéro de compte
            label: Libellé de l'écriture
            debit: Montant débit
            credit: Montant crédit
            
        Returns:
            Dictionnaire représentant l'écriture
        """
        entry = {
            "date": date_val,
            "journal": journal,
            "account": account,
            "label": label,
            "debit": float(debit) if debit else 0.0,
            "credit": float(credit) if credit else 0.0
        }
        self._entries.append(entry)
        return entry
    
    def get_entries(self) -> List[dict]:
        """
        Récupère toutes les écritures en mémoire.
        
        Returns:
            Liste des écritures
        """
        return self._entries
    
    def clear_entries(self) -> None:
        """Vide la mémoire des écritures."""
        self._entries = []
    
    def save(self, path: Optional[str] = None) -> str:
        """
        Sauvegarde le classeur Excel.
        
        Args:
            path: Chemin optionnel (utilise le chemin initial si non spécifié)
            
        Returns:
            Chemin du fichier sauvegardé
        """
        save_path = path if path else self.filename
        self.workbook.save(save_path)
        return save_path
    
    def close(self) -> None:
        """Ferme proprement le classeur (sauvegarde automatique)."""
        self.save()
    
    @property
    def entries_count(self) -> int:
        """Retourne le nombre d'écritures en mémoire."""
        return len(self._entries)
    
    @property
    def total_debit(self) -> float:
        """Calcule le total des débits."""
        return sum(e["debit"] for e in self._entries)
    
    @property
    def total_credit(self) -> float:
        """Calcule le total des crédits."""
        return sum(e["credit"] for e in self._entries)
    
    @property
    def is_balanced(self) -> bool:
        """Vérifie si les écritures sont équilibrées (Débit = Crédit)."""
        # Comparaison avec tolérance pour les flottants
        return abs(self.total_debit - self.total_credit) < 0.01
